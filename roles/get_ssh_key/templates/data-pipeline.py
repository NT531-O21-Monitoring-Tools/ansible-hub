import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os

base_path = ''
def outputxlsx():

    data = []

    current_dir = os.path.dirname(os.path.abspath(__file__))
    files = os.listdir(current_dir)
    print(files)

    values = []
    for file in files:
        file_path = os.path.join(current_dir, file)  
        if os.path.isfile(file_path) and not ('.' in file):  
            split_names = file.split('+')
            with open(file_path, 'r') as f: 
                for line in f:
                    cleaned_line = line.strip()
                    if cleaned_line:
                        values.append(cleaned_line)
            for value in values:
                data.append([split_names[0], split_names[1], split_names[2], value, ' '])

    

        

    df = pd.DataFrame(data, columns=['Region', 'Scope', 'Host name', 'SSH key', 'Describe'])
    df = df.drop_duplicates()

    with open(current_dir+'/.ignore', 'r') as f:
        ignore_namespaces = [line.strip() for line in f.readlines()]
    file_name = current_dir+"/output.xlsx"
    df = df.sort_values(by=['Region'], ascending=[True])
    df.to_excel(file_name, sheet_name='SSH', index=False)
    workbook = load_workbook(file_name)
    sheet = workbook['SSH']
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length+2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    for row in sheet.rows:
        sheet.row_dimensions[row[0].row].height = 15

    workbook.save(file_name)

if __name__ == "__main__":
    outputxlsx()